from __future__ import division
import math
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook

# code b-m for trials using only one machine

# load the workbook named DDAF_original.xlsx
DDAF_original = load_workbook('DDAF_original.xlsx')

# source of original data worksheet named original
original = DDAF_original['original']

# ############################  block-machine  ###########################
# ########## right(1)/wrong(0.5) match divided by no. block(s) ###########

# output worksheet named b-m-temp
bm = DDAF_original['b-m-temp']

def getOneBmValue(trialValue):
    machine = trialValue[:3]
    blocks = trialValue[5:]
    lob = blocks.split("&")  # list of blocks
    nob = len(lob)           # number of blocks
    bmMatch = 0              # initiate bmMatch to 0

    for block in lob:
        if machine[0].lower() == block[0].lower():
            if rule == "C":
                bmMatch += 1
            else:
                bmMatch += 0.5
        elif machine[1].upper() == block[1].upper():
            if rule == "S":
                bmMatch += 1
            else:
                bmMatch += 0.5
        else:
            bmMatch = bmMatch

    bmAverage = bmMatch / nob
    return bmAverage

def getMultiBmValue(trial):
    if trial.value.count(':::') > 0:


    # if trial.value == None:
    #     trial = trial
    #
    # # TODO add in cases with multiple machines
    # elif trial.value.count(':::') == 0 and len(trial.value) >= 6:
    #     trial.value = getOneBmValue(trial.value)
    #
    # else:
    #     trial = trial

# checks if there's only one "::"
# recur if sperated by ":::"
for i in range(2, 35):
    ruleCell = bm.cell(row=i, column=2)
    rule = ruleCell.value[1]

    for j in range(4, 80):
        thisTrial = bm.cell(row=i, column=j)
        getBmValue(thisTrial)

        print(i, j)

# save the updated workbook
DDAF_original.save('DDAF_original.xlsx')
