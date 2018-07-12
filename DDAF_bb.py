from __future__ import division
from itertools import combinations
import math
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook

# code b-b for trials using only one machine

# load the workbook named DDAF_original.xlsx
DDAF_original = load_workbook('DDAF_original.xlsx')

# source of original data worksheet named original
original = DDAF_original['original']

#############################  block-block  ############################
########## right(1)/wrong(0.5) match divided by no. relations ###########

# output worksheet named b-b-temp
bb = DDAF_original['b-b-temp']

def getOneBbValue(trialValue):
    machine = trialValue[:3]
    blocks = trialValue[5:]
    lob = blocks.split("&")  # list of blocks
    nob = len(lob)           # number of blocks
    noc = 1 if nob == 1 else (nob * (nob - 1) / 2)  # number of combination of blocks
    bbMatch = 0              # initiate bmMatch to 0

    comb = combinations(lob, 2)
    for blockPairs in list(comb):
        block = blockPairs[0]
        otherBlock = blockPairs[1]
        if block[0].lower() == otherBlock[0].lower():
            if rule == "C":
                bbMatch += 1
            else:
                bbMatch += 0.5
        elif block[1].upper() == otherBlock[1].upper():
            if rule == "S":
                bbMatch += 1
            else:
                bbMatch += 0.5
        else:
            bbMatch = bbMatch

    bbAverage = bbMatch / noc
    return bbAverage

def getBbValue(trial):
    if trial.value == None:
        trial = trial

    # TODO add in cases with multiple machines
    elif trial.value.count(':::') == 0 and len(trial.value) >= 6:
        trial.value = getOneBbValue(trial.value)

    else:
        trial = trial

# checks if there's only one "::"
# recur if sperated by ":::"
for i in range(2, 35):
    ruleCell = bb.cell(row=i, column=2)
    rule = ruleCell.value[1]

    for j in range(4, 80):
        thisTrial = bb.cell(row=i, column=j)
        getBbValue(thisTrial)

        print(i, j)

# save the updated workbook
DDAF_original.save('DDAF_original.xlsx')
