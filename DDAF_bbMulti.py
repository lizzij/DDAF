from __future__ import division
from itertools import combinations
import math
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook

# to calculate b-b concurrance of multiple machine trials

# load the workbook named DDAF_original.xlsx
DDAF_original = load_workbook('DDAF_original.xlsx')

# source of original data worksheet named original
original = DDAF_original['original']

#############################  block-block  ############################
########## right(1)/wrong(0.5) match divided by no. relations ###########

# output worksheet named b-b-temp
bb = DDAF_original['b-b-temp']

def getOneBbValue(trialValue):
    machine = trialValue[:3]
    blocks = trialValue[5:]
    lob = blocks.split("&")  # list of blocks
    nob = len(lob)           # number of blocks
    noc = 1 if nob == 1 else (nob * (nob - 1) / 2)  # number of combination of blocks
    bbMatch = 0              # initiate bmMatch to 0

    comb = combinations(lob, 2)
    for blockPairs in list(comb):
        block = blockPairs[0]
        otherBlock = blockPairs[1]
        if block[0].lower() == otherBlock[0].lower():
            if rule == "C":
                bbMatch += 1
            else:
                bbMatch += 0.5
        elif block[1].upper() == otherBlock[1].upper():
            if rule == "S":
                bbMatch += 1
            else:
                bbMatch += 0.5
        else:
            bbMatch = bbMatch

    bbAverage = bbMatch / noc
    return bbAverage

# get b-m value for multiple
def multiBbValue(trial):
    if trial.value == None or type(trial.value) is long or type(trial.value) is float:
        trial = trial

    elif trial.value.count(':::') > 0:
        sets = trial.value
        los = sets.split(":::") # list of sets
        nos = len(los)          # number of sets
        bbSet = 0
        for set in los:
            bbSet += getOneBbValue(set)
        bbResult = bbSet / nos
        trial.value = bbResult

    else:
        trial = trial

# checks if there's only one "::"
# recur if sperated by ":::"
for i in range(2, 35):
    ruleCell = bb.cell(row=i, column=2)
    rule = ruleCell.value[1]

    for j in range(4, 80):
        thisTrial = bb.cell(row=i, column=j)
        multiBbValue(thisTrial)

        print(i, j)

# save the updated workbook
DDAF_original.save('DDAF_original.xlsx')
