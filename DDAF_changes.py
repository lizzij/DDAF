# code the changes in neighbouring trials for all trials

from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook

# load the workbook named DDAF_original.xlsx
DDAF_original = load_workbook('DDAF_original.xlsx')

# source of original data worksheet named original
original = DDAF_original['original']

############################  changes ###################################
####################### next value - current value ######################

# output worksheet named number
changes = DDAF_original['changes']

for i in range(2, 162):
    for j in range(4, 80):
        current = changes.cell(row=i, column=j)
        next = changes.cell(row=i, column=(j+1))
        if (current.value is None or type(current.value) is str or type(current.value) is unicode) or \
           (next.value is None or type(next.value) is str or type(next.value) is unicode):
           current = current

        else:
            current.value = next.value - current.value

# save the updated workbook
DDAF_original.save('DDAF_original.xlsx')
