# install openpyxl
# download DDAF as .xlsx
# consolidate original (remove empty lines)
# add new sheets named "number", "b-b", and "b-m"
# finish "total" rows using AutoSum in Excel

# code the number for all trials
# and code b-b b-m for simple cases (1 machine & 1 block)

from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook

# load the workbook named DDAF_original.xlsx
DDAF_original = load_workbook('DDAF_original.xlsx')

# source of original data worksheet named original
original = DDAF_original['original']

############################  number ####################################
######## counts if the number of blocks is exactly 2 (1 for True) #######

# output worksheet named number
number = DDAF_original['number']

# if data contains one "&", indicating two blocks used
# then return 1, else return 0
for i in range(2, 35):
    for j in range(4, 80):
        nmb = number.cell(row=i, column=j)
        if nmb.value == None:
            nmb = nmb
        elif nmb.value.count('&') == 1:
            nmb.value = 1
        else:
            nmb.value = 0

# #############################  block-block  ############################
# ########## right(1)/wrong(0.5) match divided by no. relations ###########
# # output worksheet named bb
# bb = DDAF_original['b-b']
#
# # if data contains only one "::" and no "&", indicating only one block is used
# # then return 0
# # else ... (future TODO or hand code data)
# for i in range(2, 35):
#     for j in range(4, 80):
#         bbMatch = bb.cell(row=i, column=j)
#         if bbMatch.value == None:
#             bbMatch = bbMatch
#         elif bbMatch.value.count('::') == 1 and bbMatch.value.count('&') == 0:
#             bbMatch.value = 0
#         else:
#             bbMatch = bbMatch
#
# ############################  block-machine  ###########################
# ########## right(1)/wrong(0.5) match divided by no. block(s) ###########
# # output worksheet named bm
# bm = DDAF_original['b-m']
#
# # if data contains only one "::" and no "&", indicating only one block is used, then
# #   if the 1st and the 6th character matches, then
# #       if color rule C return 1, else return 0.5
# #   else if the 2nd and the 7th character matches, then
# #       if shape rule S return 1, else return 0.5
# #   else return 0
# # else ... (future TODO or hand code data)
#
# for i in range(2, 35):
#     for j in range(4, 80):
#
#         ruleCell = bm.cell(row=i, column=2)
#         rule = ruleCell.value[1]
#         bmMatch = bm.cell(row=i, column=j)
#
#         if bmMatch.value == None:
#             bmMatch = bmMatch
#
#         elif bmMatch.value.count('::') == 1 and bmMatch.value.count('&') == 0 and len(bmMatch.value) >= 7:
#             str = bmMatch.value
#             if str[0].lower() == str[5].lower():
#                 if rule == "C":
#                     bmMatch.value = 1
#                 else:
#                     bmMatch.value = 0.5
#             elif str[1].upper() == str[6].upper():
#                 if rule == "S":
#                     bmMatch.value = 1
#                 else:
#                     bmMatch.value = 0.5
#             else:
#                 bmMatch.value = 0
#
#         else:
#             bmMatch = bmMatch

# save the updated workbook
DDAF_original.save('DDAF_original.xlsx')
